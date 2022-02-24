import logging
import os
import pathlib

import pandas as pd
from openpyxl import load_workbook

from common import *


logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)



def clean_excel_files(excel_sucio_ruta):

  xls = load_workbook(filename=excel_sucio_ruta)
  sheet_list = xls.sheetnames
  df_ordered = pd.DataFrame()

  for i in range(1, len(sheet_list)):
    df_temporary = pd.read_excel(
        excel_sucio_ruta, 
        sheet_name= sheet_list[i],
        skiprows= 5
    )
    df_temporary['SHEET'] = sheet_list[i]
    df_ordered = df_ordered.append(df_temporary)

  df_ordered = df_ordered.dropna(how='any')

  return df_ordered


def merge_csv_files():
  pass



@tiempo_de_ejecucion
def run():

  cities = list(config()['municipios'].keys())

  for city in cities:
    
    CURRENT_DIR = pathlib.Path().resolve()

    BASE_PATH = CURRENT_DIR.joinpath(f"archivos-{city}", "years")
    years_folder = [año.name for año in BASE_PATH.iterdir() if año.is_dir()]

    
    for year in years_folder:
        YEAR_FILES_PATH = BASE_PATH.joinpath(str(year), "archivos")

        xlsx_files = [file.name for file in YEAR_FILES_PATH.iterdir() if file.is_file()]
        PATH_XLSX_FILES = [file for file in YEAR_FILES_PATH.iterdir() if file.is_file()]

        for file,path_file in zip(xlsx_files, PATH_XLSX_FILES):

            my_df = clean_excel_files(excel_sucio_ruta=path_file)
            my_df.to_csv(
              str(YEAR_FILES_PATH)+str(f'/{file}.csv'), index=False
            )
            
            if path_file.exists():
              path_file.unlink()



if __name__ == '__main__':
    run()
